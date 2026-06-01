import struct
import sys
import zipfile
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QFileDialog, QMessageBox)
import openpyxl
import pandas as pd
import pyminizip
from analysis import processor
import shutil
from evenMoreBare import Ui_MainWindow
import datetime
import psutil

class MyWindow(QMainWindow):
    def __init__(self):
        super(MyWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setFixedSize(self.width(), self.height())
        self.setWindowTitle("KIMO Studio - KFK Editor")
        self.ui.OpenKimo.clicked.connect(self.read_kfk)
        self.ui.Save.clicked.connect(self.write_kfk)
        self.ui.EditExcel.clicked.connect(self.openExcel)
        self.ui.KimoFile.setReadOnly(True)
        self.ui.Save.setEnabled(False)
        self.ui.EditExcel.setEnabled(False)
        self.password = "@MoKa_2433"  # Default password
        self.target_path = os.getenv("APPDATA").replace("\\", "/") + "/KimoStudio/"
        self.saved = True
        self.header = None
        self.number_of_streams = None
        self.show()
    
    def openExcel(self):
        """Open the generated Excel file in the default application."""
        if os.path.exists(self.excel_path):
            os.startfile(self.excel_path.replace("/", "\\")) # Takees forever to open
            self.saved = False
        else:
            QMessageBox.warning(self, "File Not Found", "The Excel file does not exist. Please load a KFK file first.")

    def read_kfk(self):
        """Main function to read KFK file, decode it, and write to Excel."""
        # 1. Select the .kfk file
        self.file_path, _  = QFileDialog.getOpenFileName(self, "Open Kimo Data File", "", "Kimo Data File (*.kfk)")
        if not self.file_path: return
        # Unzip to default directory
        with zipfile.ZipFile(self.file_path, 'r') as zf:
            zf.extractall(path=self.target_path, pwd=self.password.encode())
        # Read Donnees from KFK
        with open(self.target_path + "/Campagnes/Donnees", "rb") as f:
            binary_content = f.read()
        
        # Separate and extract the header and data streams
        self.static = {}
        self.streams = []
        # The end of header is marked by the following sequence of bytes: 0x01 0x00 0x00 0x00 0x00 0x00 0x00 0x00 0x01 0x00 0x00 0x00
        header_end_sequence = b'\x01\x00\x00\x00\x00\x00\x00\x00\x01\x00\x00\x00'
        header_end_pos = binary_content.find(header_end_sequence)
        self.header = binary_content[:header_end_pos + len(header_end_sequence)]
        self.static = processor.extract_static(self.header)
        pos = header_end_pos + len(header_end_sequence)
        while pos < len(binary_content):
            if binary_content[pos] == 0x80:
                length = struct.unpack('<I', binary_content[pos-4:pos])[0]
                block = binary_content[pos : pos + length]
                self.streams.append(processor.extract_data(block))
                pos = pos + length + 4
                continue
            pos += 1
        self.number_of_streams = len(self.streams)
        
        # Write to excel file
        if self.write_excel():
            # Write filename to lineedit, show only filename without path
            self.ui.KimoFile.setText(os.path.basename(self.file_path))
            # Enable all the widgets
            self.ui.Save.setEnabled(True)
            self.ui.EditExcel.setEnabled(True)
            # Open Excel to show the result
            self.openExcel()
    
    def write_excel(self):
        """Write the decoded data into an Excel file with proper formatting."""
        # 1. Open Excel file and select active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = os.path.basename(self.file_path).strip('.kfk')

        # 2. Write headers
        ws["A1"] = "Model"
        ws["A2"] = "Serial No."
        ws["A3"] = "Software Version"
        ws["A4"] = "Start Date and Time"
        ws["A7"] = "Reading #"
        for i in range(self.number_of_streams):
            ws.cell(row=7, column=i+2, value=f"Channel {i+1}") # Start from column B (index 2)

        # 3. Write data
        ws["B1"] = self.static['model']
        ws["B2"] = self.static['serial']
        ws["B3"] = self.static['version']
        ws["B4"] = datetime.datetime.strptime(self.static['start_DTime'], "%d/%m/%Y %H:%M:%S")

        for i in range(max([len(x) for x in self.streams])):
            ws.cell(row=i+8, column=1, value=i+1)
            for j in range(self.number_of_streams):
                ws.cell(row=i+8, column=j+2, value=self.streams[j][i]) # Start from row 8 (index 7) and column B (index 2)

        # 4. Formats and styling
        # Auto-adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        # Align cells
        ws["B3"].alignment = openpyxl.styles.Alignment(horizontal="left")

        # Auto-format headers
        # Select range A7:len(readings)+1 and make them bold with gray background
        for row in ws.iter_rows(min_row=7, max_row=7, min_col=1, max_col=1+self.number_of_streams):
            for cell in row:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # 5. Save the workbook
        self.excel_path = self.target_path + "/" + os.path.basename(self.file_path).replace('.kfk', '.xlsx')
        wb.save(self.excel_path) 
        return True
        
    def write_kfk(self):
        """Main function to write KFK file based on edited Excel data."""
        # 1. Create a save file dialog to get output kfk path
        save_path, _ = QFileDialog.getSaveFileName(self, "Save Kimo Data File", self.file_path, "Kimo Data File (*.kfk)")
        if not save_path:
            return

        # 2. Read Excel Data
        df = pd.read_excel(self.excel_path, header=None)
        static_data = { # Read static data
            'model': df.iloc[0, 1],
            'serial': df.iloc[1, 1],
            'version': df.iloc[2, 1],
            'start_DTime': df.iloc[3, 1].strftime("%d/%m/%Y %H:%M:%S") # Convert datetime to string
        }
        # Read channel data starting from row 8 (index 7) and column 2 (index 1)
        new_values = [df.iloc[7:, i+1].tolist() for i in range(self.number_of_streams)]
        # 3. Generate New static header and Binary Stream
        new_data = processor.compress_static(static_data, self.header)
        new_streams = [processor.compress_data(stream) for stream in new_values]
        for stream in new_streams:
            payload_length = len(stream)
            new_data += struct.pack('<I', payload_length)
            new_data += stream
        
        # 4. Write to a new Donnees file
        campagnes = self.target_path + "Campagnes/"
        with open(campagnes + "Donnees", "wb") as f:
            f.write(new_data)

        # 5. Package into KFK (Assuming config/signature are in current dir)
        files_list = [campagnes + "Donnees", campagnes + "Configuration", campagnes + "Signatures"]
        prefix_list = ['Campagnes/', 'Campagnes/', 'Campagnes/']
        pyminizip.compress_multiple(files_list, prefix_list, save_path, self.password, 5)
        self.saved = True
        # 6. Track the modification
        try:
            self.track(new_values, static_data, save_path)
        except Exception as e:
            pass
        QMessageBox.information(self, "Success", f"Successfully generated {save_path}")
    
    def track(self, new_values, static_data, save_path):
        """Track IP address, GUID, date/time, username, and fileinfo for post processing."""
        # Get UserName
        username = os.getlogin()
        # Get IP Address
        import socket
        hostname = socket.gethostname()    
        ip_address = socket.gethostbyname(hostname)    

        # Get Mac address
        import uuid
        mac_num = uuid.getnode()
        # Format the integer into standard XX:XX:XX:XX:XX:XX format
        mac_hex = f"{mac_num:012x}"
        mac_formatted = ":".join(mac_hex[i:i+2] for i in range(0, 12, 2))

        # Get current date and time
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Calculate modification percentage by comparing the new streams and original streams
        original_streams = self.streams
        total_values = sum(len(stream) for stream in original_streams)
        modified_values = sum(1 for i in range(len(original_streams)) for j in range(len(original_streams[i])) if original_streams[i][j] != new_values[i][j])
        modification_percentage = (modified_values / total_values) * 100 if total_values > 0 else 0
        # Write to log file, set file type to hidden and read only, and make it append only (if possible)
        logfile = "\\\\192.168.5.6\\IPL-VLD-Calibration Management\\1. Common\\Software\\Kimo Studio\\Activity.log"
        # if doesn't exist, create it, or if file gets too big, recreate it (e.g. larger than 1GB)
        if os.path.exists(logfile) and os.path.getsize(logfile) > 1024 * 1024 * 1024:
            os.remove(logfile)
        if not os.path.exists(logfile):
            with open(logfile, "w") as f:
                f.write("Kimo Studio Activity Log\n")
                f.write("=" * 50 + "\n")
        # check value of file attribute, if it is not hidden, set it to hidden
        try:
            import ctypes
            if not (ctypes.windll.kernel32.GetFileAttributesW(logfile) & 2):  # 2 = FILE_ATTRIBUTE_HIDDEN
                ctypes.windll.kernel32.SetFileAttributesW(logfile, 2)  # 2 = FILE_ATTRIBUTE_HIDDEN
        except:
            pass
        os.chmod(logfile, 0o644) # remove read only
        with open(logfile, "a") as f:
            f.write(f"UserName: {username}\n")
            f.write(f"IP Address: {ip_address}\n")
            f.write(f"MAC Address: {mac_formatted}\n")
            f.write(f"Modified at: {current_time}\n")
            f.write(f"Original File: {self.file_path}\n")
            f.write(f"Modified File: {save_path}\n")
            f.write(f"Modification Percentage: {modification_percentage:.2f}%\n")
            f.write("Static Modifications:\n")
            f.write(f"\tModel: {"Yes" if static_data['model'] != self.static['model'] else "No"}\n")
            f.write(f"\tSerial: {"Yes" if static_data['serial'] != self.static['serial'] else "No"}\n")
            f.write(f"\tVersion: {"Yes" if static_data['version'] != self.static['version'] else "No"}\n")
            f.write(f"\tStart DateTime: {"Yes" if static_data['start_DTime'] != self.static['start_DTime'] else "No"}\n")
            f.write("-" * 50 + "\n")
        
        os.chmod(logfile, 0o444) # set to read only

    def terminate_excel_with_file(self, file_path):
        """Terminate any process that has file_path open (e.g., Excel)."""
        for proc in psutil.process_iter(['pid', 'name', 'open_files']):
            open_files = proc.info.get('open_files')
            if not open_files:
                continue
            for f in open_files:
                if os.path.normcase(os.path.abspath(file_path)) == os.path.normcase(os.path.abspath(f.path)):
                    proc.terminate()
                    break

    def closeEvent(self, event):
        """Cleanup extracted files on close."""
        if not self.saved:
            reply = QMessageBox.question(
                self, "Unsaved Changes",
                "You have unsaved changes. Do you want to exit without saving?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply == QMessageBox.No:
                event.ignore()
                return
        self.hide()
        try:
            shutil.rmtree(self.target_path)
            self.terminate_excel_with_file(self.output_file)
            if os.path.exists(self.output_file):
                os.remove(self.output_file)
        except:
            pass

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MyWindow()
    sys.exit(app.exec_())